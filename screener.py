#!/usr/bin/env python3
from __future__ import annotations
"""
=============================================================
  INVESTMENT SCREENER AGENT v1.0


  Universe:   S&P 500 + Nasdaq 100 + Dow Jones 30
  Hard filters (4):
    1. Analyst upside to mean target  > 20%
    2. Number of analysts             >= 5
    3. Target dispersion (range/mean) < 40%
    4. Revision momentum              positive vs ~60d ago
       (skipped on first run — no history yet)

  Scoring (+1 per condition, max 3):
    +1  Days-to-cover  > 5
    +1  Put/Call OI    < 0.70
    +1  Call volume/OI > 1.50  (proxy for unusual call activity)

  AI narratives:  Gemini 2.0 Flash (free tier, score >= 1 only)
  Output:         Excel, one tab per universe, sorted score + upside
  Email:          SendGrid free tier, Excel attached
  Scheduler:      Mac cron — see setup_cron.sh

  SETUP:
    pip install yfinance pandas openpyxl sendgrid \
                google-generativeai tqdm requests

  API KEYS: set the four variables in the CONFIG section below.
=============================================================
"""

# ── CONFIG ────────────────────────────────────────────────────
# Nessun segreto nel codice. Le chiavi si leggono da variabili d'ambiente
# (vedi .env.example). In locale: copia .env.example in .env e compilalo.
import os

try:
    from dotenv import load_dotenv   # opzionale: pip install python-dotenv
    load_dotenv()
except ImportError:
    pass

GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]          # aistudio.google.com
RESEND_API_KEY = os.environ["RESEND_API_KEY"]          # resend.com → API Keys
RESEND_FROM    = os.environ.get("RESEND_FROM", "onboarding@resend.dev")
EMAIL_TO       = os.environ.get("EMAIL_TO", "")        # recipient list

# Hard filter thresholds
UPSIDE_MIN       = 0.20    # 20% minimum upside to analyst mean target
ANALYST_MIN      = 5       # minimum number of analysts covering the stock
DISPERSION_MAX   = 0.40    # (targetHigh - targetLow) / targetMean < 40%

# Score thresholds
DAYS_TO_COVER_MIN = 5.0    # short ratio (days to cover) > 5  →  +1
PC_RATIO_MAX      = 0.70   # put/call open interest ratio < 0.7  →  +1
CALL_ACTIVITY_MIN = 1.50   # call volume / call OI > 1.5  →  +1 (unusual call activity proxy)

MAX_WORKERS      = 8
DATA_STORE_PATH  = "data_store.json"   # stores historical targets for revision momentum

# ─────────────────────────────────────────────────────────────

import warnings, json, os, time
warnings.filterwarnings("ignore")

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm


# ── TICKER LISTS ──────────────────────────────────────────────

def get_sp500() -> list[str]:
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies")
        return list(tables[0]["Symbol"].str.replace(".", "-", regex=False))
    except Exception as e:
        print(f"  [WARN] SP500 fetch failed: {e}")
        return []

def get_nasdaq100() -> list[str]:
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/Nasdaq-100")
        for t in tables:
            for col in ["Ticker", "Symbol"]:
                if col in t.columns:
                    return list(t[col].dropna().str.replace(".", "-", regex=False))
    except Exception as e:
        print(f"  [WARN] Nasdaq100 fetch failed: {e}")
    return []

DOW30 = [
    "AAPL","AMGN","AXP","BA","CAT","CRM","CSCO","CVX","DIS","DOW",
    "GS","HD","HON","IBM","INTC","JNJ","JPM","KO","MCD","MMM",
    "MRK","MSFT","NKE","PG","TRV","UNH","V","VZ","WBA","WMT"
]


# ── HISTORICAL STORE (revision momentum) ──────────────────────

def load_store() -> dict:
    if os.path.exists(DATA_STORE_PATH):
        with open(DATA_STORE_PATH) as f:
            return json.load(f)
    return {}

def save_store(store: dict):
    with open(DATA_STORE_PATH, "w") as f:
        json.dump(store, f, indent=2)

def update_store(store: dict, ticker: str, target_mean: float):
    today = datetime.today().strftime("%Y-%m-%d")
    store.setdefault(ticker, {})[today] = target_mean

def revision_momentum(store: dict, ticker: str, current_target: float):
    """
    Returns:
      True  → targets rising (passes filter)
      None  → no history yet (passes filter, first run)
      False → targets falling (filtered out)
    """
    history = store.get(ticker, {})
    if len(history) < 2:
        return None
    dates = sorted(history.keys())
    cutoff = (datetime.today() - timedelta(days=60)).strftime("%Y-%m-%d")
    old_dates = [d for d in dates if d <= cutoff]
    if not old_dates:
        return None
    old_target = history[old_dates[-1]]
    return current_target > old_target


# ── PHASE 1: FUNDAMENTAL HARD FILTERS ─────────────────────────

def fetch_fundamental(ticker: str, store: dict) -> dict | None:
    """Fetch analyst data; returns a row dict if all hard filters pass, else None."""
    try:
        info = yf.Ticker(ticker).info

        target_mean = info.get("targetMeanPrice")
        target_high = info.get("targetHighPrice")
        target_low  = info.get("targetLowPrice")
        current     = info.get("currentPrice") or info.get("regularMarketPrice")
        n_analysts  = info.get("numberOfAnalystOpinions") or 0
        short_ratio = info.get("shortRatio")        # days-to-cover
        sector      = info.get("sector", "Unknown")
        name        = info.get("shortName", ticker)
        forward_pe  = info.get("forwardPE")

        # Guard: must have all core values
        if not all([target_mean, target_high, target_low, current]) or n_analysts == 0:
            return None

        # Filter 1 — upside
        upside = (target_mean - current) / current
        if upside < UPSIDE_MIN:
            return None

        # Filter 2 — analyst count
        if n_analysts < ANALYST_MIN:
            return None

        # Filter 3 — dispersion
        dispersion = (target_high - target_low) / target_mean
        if dispersion > DISPERSION_MAX:
            return None

        # Filter 4 — revision momentum (update store first, then evaluate)
        update_store(store, ticker, target_mean)
        momentum = revision_momentum(store, ticker, target_mean)
        if momentum is False:
            return None  # targets actively declining → skip

        return {
            "ticker":      ticker,
            "name":        name,
            "sector":      sector,
            "current":     round(current, 2),
            "target_mean": round(target_mean, 2),
            "target_high": round(target_high, 2),
            "target_low":  round(target_low, 2),
            "upside_pct":  round(upside * 100, 1),
            "n_analysts":  n_analysts,
            "dispersion":  round(dispersion * 100, 1),
            "days_to_cover": round(short_ratio, 1) if short_ratio else None,
            "forward_pe":  round(forward_pe, 1) if forward_pe else None,
            "revision_momentum": momentum,  # True or None
        }

    except Exception:
        return None


# ── PHASE 2: OPTIONS / UOA ────────────────────────────────────

def fetch_uoa(ticker: str) -> tuple:
    """
    Returns (pc_ratio, call_activity_ratio, call_volume) from nearest expiry.

    put/call OI ratio < 0.7          → bullish options positioning
    call volume / call OI > 1.5      → unusual call activity (proxy)

    Note: true 20d average call volume would require storing historical options
    data across runs. call_volume/OI is a session-level proxy for the same idea.
    """
    try:
        t = yf.Ticker(ticker)
        exps = t.options
        if not exps:
            return None, None, None
        chain  = t.option_chain(exps[0])
        calls  = chain.calls
        puts   = chain.puts
        c_oi   = calls["openInterest"].fillna(0).sum()
        p_oi   = puts["openInterest"].fillna(0).sum()
        c_vol  = calls["volume"].fillna(0).sum()
        pc     = round(p_oi / c_oi, 3)  if c_oi > 0 else None
        ca     = round(c_vol / c_oi, 2) if c_oi > 0 else None
        return pc, ca, int(c_vol)
    except Exception:
        return None, None, None


# ── SCORING ───────────────────────────────────────────────────

def score_stock(row: dict) -> dict:
    score   = 0
    reasons = []

    dtc = row.get("days_to_cover")
    if dtc and dtc > DAYS_TO_COVER_MIN:
        score += 1
        reasons.append(f"days-to-cover {dtc}")

    pc = row.get("pc_ratio")
    if pc and pc < PC_RATIO_MAX:
        score += 1
        reasons.append(f"P/C OI {pc}")

    ca = row.get("call_activity")
    if ca and ca > CALL_ACTIVITY_MIN:
        score += 1
        reasons.append(f"call activity {ca}×")

    row["score"]         = score
    row["score_reasons"] = ", ".join(reasons) if reasons else "—"
    return row


# ── AI NARRATIVE (Gemini 2.0 Flash) ──────────────────────────

def generate_narratives(rows: list[dict]) -> list[dict]:
    """
    Generates a 2-sentence plain-English summary for each stock with score >= 1.
    Stocks with score 0 get a placeholder to avoid unnecessary API calls.
    """
    try:
        from google import genai
        client = genai.Client(api_key=GEMINI_API_KEY)
    except ImportError:
        print("  [WARN] google-generativeai not installed — skipping narratives")
        for r in rows:
            r["narrative"] = "Install google-generativeai to enable narratives."
        return rows
    except Exception as e:
        print(f"  [WARN] Gemini init failed: {e}")
        for r in rows:
            r["narrative"] = "Narrative unavailable."
        return rows

    for row in rows:
        if row.get("score", 0) == 0:
            row["narrative"] = "Score 0 — no narrative generated."
            continue
        prompt = (
            f"You are a sell-side analyst writing a brief stock note. "
            f"In exactly 2 sentences (max 60 words total), explain why {row['name']} "
            f"({row['ticker']}, {row['sector']}) appears in a quantitative screener shortlist. "
            f"Use only these facts — no added speculation:\n"
            f"• Price ${row['current']} vs analyst mean target ${row['target_mean']} "
            f"({row['upside_pct']}% upside, {row['n_analysts']} analysts)\n"
            f"• Target range ${row['target_low']}–${row['target_high']} "
            f"(dispersion {row['dispersion']}%)\n"
            f"• Revision momentum: {'positive — targets rising' if row['revision_momentum'] else 'neutral (first run, no prior data)'}\n"
            f"• Days-to-cover: {row.get('days_to_cover') or 'n/a'}\n"
            f"• Score bonuses: {row.get('score_reasons', '—')}\n"
            f"Write in English. Factual. No hype."
        )
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash", contents=prompt
            )
            row["narrative"] = response.text.strip()
            time.sleep(0.5)  # respect free-tier rate limit (15 RPM)
        except Exception as e:
            row["narrative"] = f"Narrative error: {e}"

    return rows


# ── EXCEL OUTPUT ──────────────────────────────────────────────

COLUMN_ORDER = [
    "ticker", "name", "sector", "current", "target_mean", "upside_pct",
    "n_analysts", "dispersion", "days_to_cover", "forward_pe",
    "pc_ratio", "call_activity", "score", "score_reasons", "narrative",
]

COLUMN_LABELS = {
    "ticker":        "Ticker",
    "name":          "Name",
    "sector":        "Sector",
    "current":       "Price ($)",
    "target_mean":   "Target Mean ($)",
    "upside_pct":    "Upside (%)",
    "n_analysts":    "Analysts (#)",
    "dispersion":    "Dispersion (%)",
    "days_to_cover": "Days-to-Cover",
    "forward_pe":    "Fwd P/E",
    "pc_ratio":      "P/C OI Ratio",
    "call_activity": "Call Activity (×)",
    "score":         "Score (0–3)",
    "score_reasons": "Score Breakdown",
    "narrative":     "AI Narrative",
}

def write_excel(results_by_universe: dict, path: str):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="1F2D3D")
    HEADER_FONT  = Font(color="FFFFFF", bold=True)
    SCORE3_FILL  = PatternFill("solid", fgColor="D4EFDF")  # light green
    SCORE2_FILL  = PatternFill("solid", fgColor="FEF9E7")  # light yellow

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for universe, rows in results_by_universe.items():
            if not rows:
                pd.DataFrame([{"note": "No stocks passed filters this week."}]).to_excel(
                    writer, sheet_name=universe, index=False)
                continue

            df = pd.DataFrame(rows)
            cols = [c for c in COLUMN_ORDER if c in df.columns]
            df   = df[cols].rename(columns=COLUMN_LABELS)
            df   = df.sort_values(
                [COLUMN_LABELS["score"], COLUMN_LABELS["upside_pct"]],
                ascending=[False, False]
            ).reset_index(drop=True)

            df.to_excel(writer, sheet_name=universe, index=False)
            ws = writer.sheets[universe]

            # Header styling
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            # Row highlighting by score
            score_col_idx = df.columns.get_loc(COLUMN_LABELS["score"]) + 1
            for row_idx, row_data in enumerate(ws.iter_rows(min_row=2), start=2):
                score_val = ws.cell(row=row_idx, column=score_col_idx).value
                fill = (SCORE3_FILL if score_val == 3 else
                        SCORE2_FILL if score_val == 2 else None)
                if fill:
                    for cell in row_data:
                        cell.fill = fill

            # Column widths
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

    print(f"  ✓ Excel saved: {path}")


# ── EMAIL (SendGrid) ──────────────────────────────────────────

def send_email(excel_path: str, results_by_universe: dict):
    try:
        import httpx, base64

        with open(excel_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()

        rows_html = "".join(
            f"<tr><td><b>{u}</b></td><td>{len(r)} stocks</td>"
            f"<td>{sum(1 for x in r if x.get('score',0)>=2)} con score ≥2</td></tr>"
            for u, r in results_by_universe.items()
        )

        payload = {
            "from": RESEND_FROM,
            "to": EMAIL_TO,
            "subject": f"📊 Weekly Screener — {datetime.today().strftime('%d %b %Y')}",
            "html": f"""
<html><body style="font-family:Arial,sans-serif;color:#222;">
<h2 style="color:#1F2D3D;">📊 Weekly Stock Screener — {datetime.today().strftime('%d %b %Y')}</h2>
<p>Shortlist settimanale basata su upside analitico, dispersione target, revision momentum
e options flow. Filtri: upside &gt;20%, ≥5 analisti, bassa dispersione, momentum positivo.</p>
<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;">
  <tr style="background:#1F2D3D;color:#fff;">
    <th>Universe</th><th>Stocks shortlisted</th><th>High conviction (score ≥2)</th>
  </tr>
  {rows_html}
</table>
<p style="margin-top:16px;">Report completo con AI narratives allegato in Excel.<br>
<small>🟢 Righe verdi = score 3 | 🟡 Righe gialle = score 2</small></p>
</body></html>""",
            "attachments": [{
                "filename": os.path.basename(excel_path),
                "content": encoded,
            }]
        }

        response = httpx.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
            json=payload,
            timeout=30
        )
        response.raise_for_status()
        print(f"  ✓ Email inviata via Resend (id: {response.json().get('id')})")

    except Exception as e:
        print(f"  [WARN] Email fallita: {e}")


# ── MAIN ORCHESTRATOR ─────────────────────────────────────────

def screen_universe(tickers: list, label: str, store: dict) -> list[dict]:
    print(f"\n{'─'*52}")
    print(f"  {label}  ({len(tickers)} tickers)")
    print(f"{'─'*52}")

    # Phase 1: hard filters — parallel
    passed = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fetch_fundamental, t, store): t for t in tickers}
        for fut in tqdm(as_completed(futures), total=len(futures),
                        desc=f"  [{label}] fundamentals"):
            result = fut.result()
            if result:
                passed.append(result)

    print(f"  → {len(passed)} passed hard filters")

    if not passed:
        return []

    # Phase 2: UOA — only on shortlisted stocks (much smaller set)
    results = []
    for row in tqdm(passed, desc=f"  [{label}] options/UOA"):
        pc, ca, cv       = fetch_uoa(row["ticker"])
        row["pc_ratio"]  = pc
        row["call_activity"] = ca
        row["call_volume"]   = cv
        results.append(score_stock(row))

    return results


def main():
    print("\n" + "=" * 52)
    print("  INVESTMENT SCREENER AGENT v1.0")
    print(f"  {datetime.today().strftime('%A %d %B %Y  %H:%M')}")
    print("=" * 52)

    store = load_store()

    universes = {
        "S&P 500":    get_sp500(),
        "Nasdaq 100": get_nasdaq100(),
        "Dow Jones":  DOW30,
    }

    results_by_universe = {}
    for label, tickers in universes.items():
        if not tickers:
            print(f"  [WARN] No tickers for {label}, skipping.")
            results_by_universe[label] = []
            continue
        results_by_universe[label] = screen_universe(tickers, label, store)

    save_store(store)

    # Phase 3: AI narratives (only score >= 1, to minimise free-tier API calls)
    print("\n  Generating AI narratives (score ≥ 1)...")
    for label, rows in results_by_universe.items():
        results_by_universe[label] = generate_narratives(rows)

    # Output
    date_str   = datetime.today().strftime("%Y%m%d")
    excel_path = os.path.expanduser(f"~/Desktop/screener_{date_str}.xlsx")
    write_excel(results_by_universe, excel_path)

    # Summary
    print("\n📊 SUMMARY")
    total = 0
    for label, rows in results_by_universe.items():
        s2 = sum(1 for r in rows if r.get("score", 0) >= 2)
        print(f"  {label:15s}  {len(rows):3d} shortlisted  |  {s2} with score ≥2")
        total += len(rows)
    print(f"  {'TOTAL':15s}  {total:3d} stocks\n")

    send_email(excel_path, results_by_universe)

    print("✅  Done.\n")


if __name__ == "__main__":
    main()